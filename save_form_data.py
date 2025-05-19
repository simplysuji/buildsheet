import pandas as pd
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import streamlit as st

def process_sap_form_data(form_data_path, template_path, output_path=None):
    """
    Process the SAP form data and populate the Excel template.
    
    Args:
        form_data_path (str): Path to the JSON file containing form data
        template_path (str): Path to the Excel template file
        output_path (str, optional): Path to save the filled template. If None, overwrites the template.
    
    Returns:
        str: Path to the saved Excel file
    """
    # Load the form data
    with open(form_data_path, 'r') as f:
        form_data = json.load(f)
    
    # Extract general configuration and server data
    general_config = form_data.get("general_config", {})
    server_data = form_data.get("server_data", [])
    
    # Load the Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["SAP"]  # Target the "SAP" sheet
    
    # Start filling data from row 12
    start_row = 12
    
    # For each server, create a new row in the Excel sheet
    for server_idx, server in enumerate(server_data):
        row = start_row + server_idx
        
        # Map form fields to Excel columns
        # Based on the PDF document structure, this mapping may need adjustments
        column_mapping = {
            # These are mappings based on the PDF structure
            # A: Environment (from general config)
            # B: SID (from general config)
            # C: Instance Number (from server config)
            # D: Server Role (from server config)
            # E: Service Criticality (from server config)
            # F: ITSG ID (from general config)
            # G: OS Version (from server config)
            # H: Azure Instance Type (from server config)
            # I: Memory/CPU (from server config)
            # ... additional mappings as needed
        }
        
        # Fill in the data row by row
        sheet.cell(row=row, column=1).value = general_config.get("Environment", "")  # A: Environment
        sheet.cell(row=row, column=2).value = general_config.get("SID", "")  # B: SID
        sheet.cell(row=row, column=3).value = server.get("Instance Number", "")  # C: Instance Number
        sheet.cell(row=row, column=4).value = server.get("Server Role", "")  # D: Server Role
        sheet.cell(row=row, column=5).value = server.get("Service Criticality", "")  # E: Service Criticality
        sheet.cell(row=row, column=6).value = general_config.get("ITSG ID", "")  # F: ITSG ID
        sheet.cell(row=row, column=7).value = server.get("OS Version", "")  # G: OS Version
        
        # Assuming "Licensing" is not in the form
        sheet.cell(row=row, column=8).value = "UL: BYOL"  # H: Licensing (default value)
        
        sheet.cell(row=row, column=9).value = server.get("Instance Type", "")  # I: Azure Instance Type
        sheet.cell(row=row, column=10).value = server.get("Memory/CPU", "")  # J: Memory/CPU
        
        # Other fields from the PDF that might need to be mapped:
        # K: Service Model name
        # L: A Record / CNAME
        sheet.cell(row=row, column=12).value = general_config.get("A Record / CNAME", "")
        
        # M: Subnet/Zone
        sheet.cell(row=row, column=13).value = general_config.get("Subnet/Zone", "")
        
        # N: Azure Subscription
        sheet.cell(row=row, column=14).value = general_config.get("Azure Subscription", "")
        
        # O: Azure Resource Group Name (may need a default or derived value)
        # P: Cluster
        sheet.cell(row=row, column=16).value = general_config.get("Cluster", "")
        
        # Q: Proximity Placement Group (may need a default or derived value)
        # R: Availability Set (may need a default or derived value)
        # S: Azure Availability Zone
        sheet.cell(row=row, column=19).value = general_config.get("AZ Selection", "")
        
        # T: Accelerated Networking (may need a default value)
        # U: OptInOptOut
        sheet.cell(row=row, column=21).value = general_config.get("OptInOptOut", "")
        
        # V: Park My Cloud Schedule
        sheet.cell(row=row, column=22).value = general_config.get("Park My Cloud Schedule", "") if general_config.get("OptInOptOut") == "In" else ""
        
        # W: Park My cloud team name and Member
        sheet.cell(row=row, column=23).value = general_config.get("Park My cloud team name and Member", "") if general_config.get("OptInOptOut") == "In" else ""
        
        # X: Disk Size (GB) (may need a default or derived value)
        # Y: Storage Type (may need a default or derived value)
        # Z: Outbound Internet Access Required
        sheet.cell(row=row, column=26).value = general_config.get("Outbound Internet Access Required", "")
        
        # AA: Time Zone (may need a default value or mapping)
        
    # Save the workbook
    save_path = output_path if output_path else template_path
    workbook.save(save_path)
    return save_path

def save_form_data_to_json(form_data, output_path):
    """
    Save the form data to a JSON file.
    
    Args:
        form_data (dict): Dictionary containing form data
        output_path (str): Path to save the JSON file
    
    Returns:
        str: Path to the saved JSON file
    """
    with open(output_path, 'w') as f:
        json.dump(form_data, f, indent=4)
    return output_path

# Example usage in a Streamlit app
def display_processing_page():
    st.title("Process SAP Form Data")
    
    # File uploader for JSON data
    uploaded_json = st.file_uploader("Upload Form Data (JSON)", type=["json"])
    
    # Check if Template.xlsx exists in the current directory
    template_path = "Template.xlsx"
    template_exists = os.path.exists(template_path)
    
    if not template_exists:
        st.error(f"Template file '{template_path}' not found in the current directory.")
        return
    
    st.info(f"Found template file: {template_path}")
    
    if uploaded_json is not None:
        # Save the uploaded JSON temporarily
        temp_json_path = "temp_form_data.json"
        with open(temp_json_path, "wb") as f:
            f.write(uploaded_json.getbuffer())
        
        # Output file path
        output_excel_path = "Filled_Template.xlsx"
        
        if st.button("Process Data"):
            try:
                # Process the data
                result_path = process_sap_form_data(temp_json_path, template_path, output_excel_path)
                
                st.success(f"Successfully processed form data and saved to: {result_path}")
                
                # Provide download link for the filled template
                with open(result_path, "rb") as file:
                    st.download_button(
                        label="Download Filled Template",
                        data=file,
                        file_name="SAP_Filled_Template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Error processing data: {str(e)}")
    else:
        st.warning("Please upload the form data JSON file.")

if __name__ == "__main__":
    display_processing_page()