import pandas as pd
import json
import os
import openpyxl

def generate_service_model_names(server_role, sid, sap_region, city, region_code="eu", dns_excel_path="SAP Buildsheet Automation Feeder.xlsx", aas_counter=None):
    """
    Generate service model names based on server role by reading from DNS Excel file
    
    Args:
        server_role (str): The role of the server
        sid (str): SID of the system
        sap_region (str): SAP Region (Sirius, U2K2, etc.)
        city (str): City name for the server (e.g., Amsterdam, Dublin)
        region_code (str): Region code for the URL (default: "eu")
        dns_excel_path (str): Path to the DNS naming convention Excel file
        aas_counter (int): Counter for AAS servers (1, 2, 3, etc.)
    
    Returns:
        str: Generated service model name(s)
    """
    # Map SAP Region to region code letter
    region_mapping = {
        "Sirius": "e",
        "U2K2": "a",
        "Cordillera": "t",
        "Global": "o",
        "POC/Model Env": "x"
    }
    
    # Get the region letter from mapping or default to "x"
    region_letter = region_mapping.get(sap_region, "x")
    
    # Read DNS mapping from Excel file
    try:
        dns_df = pd.read_excel(dns_excel_path, sheet_name="DNS -Azure", header=1)
        # Create a mapping from Identifier to DNS code (number + letter)
        identifier_mapping = {}
        for index, row in dns_df.iterrows():
            if pd.notna(row['Identifier']):
                # Get the DNS code from the No. column if it exists
                if 'No.' in dns_df.columns and pd.notna(row['No.']):
                    dns_code = str(row['No.']).strip()  # Keep as is (e.g., "001a")
                else:
                    # Extract the code from FQDN Example if No. column doesn't exist or is empty
                    fqdn_example = row['FQDN Example'] if pd.notna(row['FQDN Example']) else ""
                    dns_code = ""
                    if fqdn_example:
                        # Extract the code part after the SID
                        parts = fqdn_example.split("<")
                        if len(parts) >= 2 and ">" in parts[1]:
                            code_part = parts[1].split(">")[1]
                            # Extract the code (3 digits + letter)
                            import re
                            code_match = re.search(r'\d{3}[a-z]', code_part)
                            if code_match:
                                dns_code = code_match.group(0)
                
                identifier_mapping[row['Identifier']] = dns_code
    except Exception as e:
        print(f"Warning: Could not read DNS mapping from Excel. Using default mapping. Error: {str(e)}")
        # Fallback mapping in case the Excel cannot be read
        identifier_mapping = {
            "NFS": "000a",
            "HANA DB": "001a",
            "SQL DB": "001a",
            "ASCS": "002a",
            "Central": "002a",
            "PAS": "051a",
            "SCS": "003a",
            "AAS-Amsterdam": "102a",
            "AAS-Dublin": "202a"
        }
    
    # Handle AAS server numbering for multiple AAS servers
    if "AAS" in server_role and aas_counter is not None:
        # Determine base number based on city
        if city == "Amsterdam" or "Amsterdam" in city:
            base_number = 102
        elif city == "Dublin" or "Dublin" in city:
            base_number = 202
        else:
            base_number = 102  # Default to Amsterdam
        
        # Calculate the new number (102, 103, 104... or 202, 203, 204...)
        new_number = base_number + aas_counter - 1
        new_dns_code = f"{new_number:03d}a"
        
        # Update the mapping for AAS with the correct city
        for identifier in identifier_mapping:
            if "AAS" in identifier and (city in identifier or 
                ("Amsterdam" in identifier and ("Amsterdam" in city or city == "Amsterdam")) or
                ("Dublin" in identifier and ("Dublin" in city or city == "Dublin"))):
                identifier_mapping[identifier] = new_dns_code
                break
    
    # Handle combined roles (split by +)
    if "+" in server_role:
        combined_roles = server_role.split("+")
        model_names = []
        
        for role in combined_roles:
            role = role.strip()
            # Match role with identifier in mapping
            matched_identifier = None
            for identifier in identifier_mapping:
                if "AAS" in role:
                    # Special case for AAS, based on Amsterdam or Dublin
                    if ("AAS" in identifier and 
                        (city in identifier or 
                         ("Amsterdam" in identifier and ("Amsterdam" in city or city == "Amsterdam")) or
                         ("Dublin" in identifier and ("Dublin" in city or city == "Dublin")))):
                        matched_identifier = identifier
                        break
                elif identifier in role:
                    matched_identifier = identifier
                    break
            
            if matched_identifier and identifier_mapping[matched_identifier]:
                dns_code = identifier_mapping[matched_identifier]
                # Create model name with the format <region_letter><sid><dns_code>.<region_code>.unilever.com
                model_name = f"{region_letter}{sid.lower()}{dns_code}.{region_code}.unilever.com"
                model_names.append(model_name)
        
        # Join all model names with semicolons
        return "; ".join(model_names)
    else:
        # Handle single role
        matched_identifier = None
        # First, try to find an exact match (prioritize HA variants)
        for identifier in identifier_mapping:
            if "AAS" in server_role:
                # Special case for AAS, based on Amsterdam or Dublin
                if (identifier == server_role and 
                    (city in identifier or 
                    ("Amsterdam" in identifier and ("Amsterdam" in city or city == "Amsterdam")) or
                    ("Dublin" in identifier and ("Dublin" in city or city == "Dublin")))):
                    matched_identifier = identifier
                    break
            elif identifier == server_role:  # Exact match
                matched_identifier = identifier
                break

        # If no exact match found, fall back to partial matching
        if not matched_identifier:
            for identifier in identifier_mapping:
                if "AAS" in server_role:
                    # Special case for AAS, based on Amsterdam or Dublin
                    if ("AAS" in identifier and 
                        (city in identifier or 
                        ("Amsterdam" in identifier and ("Amsterdam" in city or city == "Amsterdam")) or
                        ("Dublin" in identifier and ("Dublin" in city or city == "Dublin")))):
                        matched_identifier = identifier
                        break
                elif identifier in server_role:
                    matched_identifier = identifier
                    break
        
        if matched_identifier and identifier_mapping[matched_identifier]:
            dns_code = identifier_mapping[matched_identifier]
            # Create model name with the format <region_letter><sid><dns_code>.<region_code>.unilever.com
            return f"{region_letter}{sid.lower()}{dns_code}.{region_code}.unilever.com"
    
    # Default return empty if no mapping found
    return ""

def get_instance_number(server_role, environment_type, excel_path="SAP Buildsheet Automation Feeder.xlsx"):
    """
    Get instance number for server role based on environment type by reading from Excel file
    
    Args:
        server_role (str): The server role
        environment_type (str): Environment type (Production or Non-Production)
        excel_path (str): Path to the Excel file containing instance numbers
        
    Returns:
        str: Instance number for the given server role and environment
    """
    import pandas as pd
    
    # Determine environment column to use
    env_column = "Production & DR" if "Production" in environment_type else "Non-Production"
    
    # Check if this is a clustered configuration (has -HA suffix) and is Production
    is_clustered = "-HA" in server_role and "Production" in environment_type
    
    # Read instance numbers from Excel file
    try:
        df = pd.read_excel(excel_path, sheet_name="Instance Number")
        
        # Create a mapping from Server roles to instance numbers
        instance_mapping = {}
        
        # Create mapping by iterating through rows
        for _, row in df.iterrows():
            server = row.iloc[0]  # Column A (Server Role)
            type_val = row.iloc[1]  # Column B (Type)
            nonprod_val = row.iloc[2]  # Column C (Non-Production)
            prod_val = row.iloc[3]  # Column D (Production & DR)
            
            if pd.notna(server) and str(server).strip() and pd.notna(type_val):
                server_key = str(server).strip()
                type_key = str(type_val).strip()
                
                if server_key not in instance_mapping:
                    instance_mapping[server_key] = {}
                
                # Convert to string while preserving leading zeros
                nonprod_clean = ""
                if pd.notna(nonprod_val):
                    if isinstance(nonprod_val, float) and nonprod_val.is_integer():
                        # Convert to int first to remove decimal, then format with leading zeros
                        nonprod_clean = f"{int(nonprod_val):02d}"
                    else:
                        nonprod_clean = str(nonprod_val).strip()
                
                prod_clean = ""
                if pd.notna(prod_val):
                    if isinstance(prod_val, float) and prod_val.is_integer():
                        # Convert to int first to remove decimal, then format with leading zeros
                        prod_clean = f"{int(prod_val):02d}"
                    else:
                        prod_clean = str(prod_val).strip()
                
                instance_mapping[server_key][type_key] = {
                    "Non-Production": nonprod_clean,
                    "Production & DR": prod_clean
                }
                
    except Exception as e:
        print(f"Warning: Could not read instance numbers from Excel for {server_role}. Using default mapping. Error: {str(e)}")
        # Default mapping in case of errors
        instance_mapping = {
            "SAP HANA DB": {
                "Standalone": {"Non-Production": "00", "Production & DR": "00"},
                "Clustered": {"Non-Production": "00", "Production & DR": "00"}
            },
            "SAP ASCS": {
                "Standalone": {"Non-Production": "01", "Production & DR": "21"},
                "Clustered": {"Non-Production": "01", "Production & DR": "21"}
            },
            "SAP SCS": {
                "Standalone": {"Non-Production": "02", "Production & DR": "22"},
                "Clustered": {"Non-Production": "02", "Production & DR": "22"}
            },
            "SAP PAS": {
                "Standalone": {"Non-Production": "00", "Production & DR": "20"},
                "Clustered": {"Non-Production": "00", "Production & DR": "20"}
            },
            "SAP AAS1..n": {
                "Standalone": {"Non-Production": "00", "Production & DR": "20"},
                "Clustered": {"Non-Production": "00", "Production & DR": "20"}
            },
            "SAP Web Dispatcher": {
                "Standalone": {"Non-Production": "10", "Production & DR": "20"},
                "Clustered": {"Non-Production": "10", "Production & DR": "20"}
            },
            "SAP ERS - ABAP": {
                "Clustered": {"Non-Production": "31", "Production & DR": "31"}
            },
            "SAP ERS - JAVA": {
                "Clustered": {"Non-Production": "32", "Production & DR": "32"}
            }
        }
    
    # Handle combined roles (split by +)
    if "+" in server_role:
        combined_roles = server_role.split("+")
        instance_numbers = []
        
        for role in combined_roles:
            role_clean = role.strip()
            # Remove -HA, -DR suffixes for mapping lookup
            base_role = role_clean.split("-")[0] if "-" in role_clean else role_clean
            sap_role = f"SAP {base_role}"
            
            # Find the matching entry in the mapping
            matched_role = None
            for mapped_role in instance_mapping:
                if mapped_role == sap_role or base_role.upper() in mapped_role.upper():
                    matched_role = mapped_role
                    break
            
            if matched_role:
                # Determine type (Clustered or Standalone)
                config_type = "Clustered" if is_clustered else "Standalone"
                
                if config_type in instance_mapping[matched_role] and env_column in instance_mapping[matched_role][config_type]:
                    instance_numbers.append(instance_mapping[matched_role][config_type][env_column])
        
        # Join all instance numbers with commas
        return ",".join(instance_numbers)
    else:
        # Handle single roles with suffixes like -01, -HA, -DR
        # Remove suffixes for mapping lookup
        base_role = server_role.split("-")[0] if "-" in server_role else server_role
        
        # Special handling for AAS (map to AAS1..n)
        if base_role.upper() == "AAS":
            sap_role = "SAP AAS1..n"
        else:
            sap_role = f"SAP {base_role}"
        
        # Find the matching entry in the mapping
        matched_role = None
        for mapped_role in instance_mapping:
            if mapped_role == sap_role or base_role.upper() in mapped_role.upper():
                matched_role = mapped_role
                break
        
        if matched_role:
            # Determine type (Clustered or Standalone)
            config_type = "Clustered" if is_clustered else "Standalone"
            
            # Check if the configuration type exists for this role
            if config_type in instance_mapping[matched_role] and env_column in instance_mapping[matched_role][config_type]:
                return instance_mapping[matched_role][config_type][env_column]
            elif "Standalone" in instance_mapping[matched_role] and env_column in instance_mapping[matched_role]["Standalone"]:
                # Fallback to Standalone if Clustered doesn't exist
                return instance_mapping[matched_role]["Standalone"][env_column]
    
    # Default return empty if no mapping found
    return ""

def get_sap_region_letter(sap_region):
    """Get the letter code for a SAP region"""
    region_mapping = {
        "Sirius": "e",
        "U2K2": "a",
        "Cordillera": "t",
        "Global": "o",
        "POC/Model Env": "x"
    }
    return region_mapping.get(sap_region, "x")

def get_environment_code(environment):
    """Get the code for a specific environment"""
    environment_mapping = {
        "Fix Development": "d",
        "Project Development": "d", 
        "Fix Quality": "q",
        "Project Quality": "q",
        "Training": "q",
        "Fix Performance": "q",
        "Project performance": "q",
        "Project UAT": "q",
        "Fix Regression": "u",
        "Sandbox": "s",
        "Production": "p"
    }
    return environment_mapping.get(environment, "d") # default to 'd' if not found

    
def add_load_balancer_sheet(workbook, general_config, primary_servers):
    """
    Add Azure Load Balancer sheet data based on server roles.
    
    Args:
        workbook: Excel workbook object
        general_config: General configuration dictionary
        primary_servers: List of primary server data
    """
    # Check if Azure Load Balancer sheet exists
    if "Azure Load Balancer" not in workbook.sheetnames:
        return
    
    lb_sheet = workbook["Azure Load Balancer"]
    
    
    # Check if any server has required roles
    required_roles = ["ASCS", "DB2", "HANA", "MaxDB"]
    has_required_roles = any(
        any(role in server.get("Server Role", "") for role in required_roles)
        for server in primary_servers
    )
    
    if not has_required_roles:
        return
    
    # Get common configuration values
    region_code = general_config.get("Azure Region Code", "").lower()
    environment_code = get_environment_code(general_config.get("Environment", ""))
    itsg_id = general_config.get("ITSG ID", "")
    sid = general_config.get("SID", "").upper()
    subscription = general_config.get("Azure Subscription", "")
    subscription_number = subscription.split("-")[1].split(" ")[0] if "-" in subscription else "01"
    
    # Generate resource group name
    resource_group_name = f"{region_code}-sp{subscription_number}-{environment_code}-{itsg_id}-{sid}-IS01-rg"
    
    current_row = 13
    
    # Process each server and add load balancer entries
    for server in primary_servers:
        server_role = server.get("Server Role", "")
        
        # Check which roles this server has
        has_ascs = "ASCS" in server_role
        has_db = any(db_role in server_role for db_role in ["DB2", "HANA", "MaxDB"])
        
        if not (has_ascs or has_db):
            continue
        
        # Generate service model name for this server
        azure_region = general_config.get("Azure Region", "")
        city = "Dublin"  # Default
        if "Amsterdam" in azure_region:
            city = "Amsterdam"
        elif "DR" in server_role:
            city = "Amsterdam" if "Dublin" in azure_region else "Dublin"
        
        service_model_name = generate_service_model_names(
            server.get("Server Role", ""), 
            general_config.get("SID", ""),
            general_config.get("SAP Region", "Sirius"),
            city,
            aas_counter=None
        )
        smn = service_model_name.split(".")[0]
        
        # Add ASCS load balancer if present
        if has_ascs:
            load_balancer_name = f"{region_code}-sp{subscription_number}-{environment_code}-{smn}-lb01"
            
            # Fill the row
            lb_sheet.cell(row=current_row, column=1).value = load_balancer_name
            lb_sheet.cell(row=current_row, column=2).value = "Standard"
            lb_sheet.cell(row=current_row, column=3).value = "Internal"
            lb_sheet.cell(row=current_row, column=4).value = resource_group_name
            lb_sheet.cell(row=current_row, column=5).value = "Prod"
            lb_sheet.cell(row=current_row, column=13).value = "ASCS"
            
            current_row += 1
        
        # Add DB load balancer if present
        if has_db:
            load_balancer_name = f"{region_code}-sp{subscription_number}-{environment_code}-{smn}-lb01"
            
            # Fill the row
            lb_sheet.cell(row=current_row, column=1).value = load_balancer_name
            lb_sheet.cell(row=current_row, column=2).value = "Standard"
            lb_sheet.cell(row=current_row, column=3).value = "Internal"
            lb_sheet.cell(row=current_row, column=4).value = resource_group_name
            lb_sheet.cell(row=current_row, column=5).value = "Prod"
            lb_sheet.cell(row=current_row, column=13).value = "DB"
            
            current_row += 1

def add_other_sheets(json_file_path, template_path, workbook):
    
    # Check if files exist
    if not os.path.exists(json_file_path):
        raise FileNotFoundError(f"JSON file not found: {json_file_path}")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Excel template not found: {template_path}")
    
    # Load JSON data
    with open(json_file_path, 'r') as file:
        form_data = json.load(file)
    
    # Extract general config and server data
    general_config = form_data.get("general_config", {})
    server_data = form_data.get("server_data", [])
    
    # Get environment type and basic config
    environment = general_config.get("Environment", "").lower()
    is_production = "production" in environment or "prod" in environment
    sid = general_config.get("SID", "").upper()
    sid_lower = sid.lower()
    sap_region = general_config.get("SAP Region", "Global")
    region_letter = get_sap_region_letter(sap_region)
    
    region = "AMS" if "Amsterdam" in general_config.get("Azure Region", "") else "Dublin"
    region_dr = "Dublin" if "Amsterdam" in general_config.get("Azure Region", "") else "AMS"

    # Analyze server roles and their versions
    server_roles = [server.get("Server Role", "") for server in server_data]
    
    # Create a mapping of server roles to their versions
    role_versions = {}
    for server in server_data:
        role = server.get("Server Role", "")
        version = server.get("Server Role Version", "")
        if role and role not in role_versions:
            role_versions[role] = version
    
    role_flags = {
        'has_nfs': any("+" in role for role in server_roles),
        'has_hana': any("HANA" in role for role in server_roles),
        'has_db2': any("DB2" in role for role in server_roles),
        'has_maxdb': any("MaxDB" in role for role in server_roles),
        'has_ascs_dr': any(role in ["ASCS-DR", "SCS-DR"] for role in server_roles),
        'has_ascs': any(role in ["ASCS", "SCS"] for role in server_roles),
        'has_apo': any("APO" in role for role in server_roles),
        'has_iq': any("IQ" in role for role in server_roles),
        'has_optimizer': any("Optimizer" in role for role in server_roles),
        'only_pas_aas': any(role in {"PAS", "AAS", "PAS-DR", "AAS-DR"} for role in server_roles)
    }

    # Get AFS server names
    afs_servername = next((server.get("AFS Server Name", "") for server in server_data if server.get("Server Role") == "ASCS"), None)
    afs_servername_dr = next((server.get("AFS Server Name", "") for server in server_data if server.get("Server Role") == "ASCS-DR"), None)
    
    # Collect service model names
    service_model_names = []
    for server in server_data:
        role = server.get("Server Role", "")
        server_model_name = generate_service_model_names(
            role, sid, sap_region, general_config.get("Azure Region Code", "eu")[:2]
        )
        if server_model_name:
            service_model_names.append(server_model_name)

    # =============================================================================
    # HELPER FUNCTIONS
    # =============================================================================
    
    def replace_sid_in_sheet(sheet, sid_value):
        """Replace all 'SID' occurrences with actual SID value in a sheet"""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "SID" in cell.value:
                    cell.value = cell.value.replace("SID", sid_value)
    
    def get_possible_sheet_names(base_role, environment_suffix):
        """Generate possible sheet names for a given role and environment"""
        possible_names = []
        
        # Get the version for this role
        role_version = role_versions.get(base_role, "")
        
        if is_production and environment_suffix == "Prod":
            if role_version:
                # Primary: {Server Role}_{Server Role Version} Prod
                possible_names.append(f"{base_role}_{role_version} {environment_suffix}")
            
            # Fallback: {Server Role} Prod (without version)
            possible_names.append(f"{base_role} {environment_suffix}")
            
            # Additional fallback for variations like v1, v2 if no version found
            if not role_version:
                possible_names.extend([
                    f"{base_role}_v1 {environment_suffix}",
                    f"{base_role}_v2 {environment_suffix}"
                ])
        else:
            # Non-production: just {Server Role} Non_Prod
            possible_names.append(f"{base_role} {environment_suffix}")
        
        return possible_names
    
    def select_and_rename_sheet_with_version(base_role, target_name):
        """Select the appropriate sheet based on role, version, and environment"""
        selected_sheet = None
        sheets_to_remove = []
        
        # Determine environment suffix
        env_suffix = "Prod" if is_production else "Non_Prod"
        
        # Get all possible sheet names for this role
        possible_names = get_possible_sheet_names(base_role, env_suffix)
        
        # Find all sheets that match this base role pattern
        all_role_sheets = []
        for sheet_name in workbook.sheetnames:
            if base_role.lower() in sheet_name.lower() and ("prod" in sheet_name.lower() or "non_prod" in sheet_name.lower()):
                all_role_sheets.append(sheet_name)
        
        # Try to find the best matching sheet
        for sheet_name in possible_names:
            # Check if sheet exists with case-insensitive comparison
            matching_sheet = next((name for name in workbook.sheetnames if name.lower() == sheet_name.lower()), None)
            if matching_sheet:
                selected_sheet = workbook[matching_sheet]
                break
        
        # Mark other sheets for removal
        for sheet_name in all_role_sheets:
            if selected_sheet is None or sheet_name != selected_sheet.title:
                sheets_to_remove.append(sheet_name)
        
        # Rename the selected sheet
        if selected_sheet:
            selected_sheet.title = target_name
        
        # Remove unused sheets
        for sheet_name in sheets_to_remove:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        
        return workbook[target_name] if target_name in workbook.sheetnames else None
    
    def select_and_rename_sheet(sheet_mapping, target_name):
        """Select the appropriate sheet based on environment and rename it (for simple cases)"""
        selected_sheet = None
        sheets_to_remove = []
        
        for sheet_name, sheet_env in sheet_mapping.items():
            # Check if sheet exists with case-insensitive comparison
            matching_sheet = next((name for name in workbook.sheetnames if name.lower() == sheet_name.lower()), None)
            if matching_sheet:
                if (is_production and sheet_env == "prod") or (not is_production and sheet_env == "non_prod"):
                    selected_sheet = workbook[matching_sheet]
                    selected_sheet.title = target_name
                else:
                    sheets_to_remove.append(matching_sheet)
        
        # Remove unused sheets
        for sheet_name in sheets_to_remove:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        
        return workbook[target_name] if target_name in workbook.sheetnames else None
    
    def remove_sheets_if_unused(sheet_names):
        """Remove sheets if they exist in workbook"""
        for sheet_name in sheet_names:
            # Check if sheet exists with case-insensitive comparison
            matching_sheet = next((name for name in workbook.sheetnames if name.lower() == sheet_name.lower()), None)
            if matching_sheet:
                workbook.remove(workbook[matching_sheet])

    # =============================================================================
    # CUSTOM LOGIC FUNCTIONS
    # =============================================================================
    
    def apply_nfs_logic(sheet, sid, sid_lower, region_letter, service_model_names):
        """Apply NFS-specific cell updates"""
        # A2: Combined server names with '/' separator
        if service_model_names:
            sheet.cell(row=2, column=1).value = f"{region_letter}{sid_lower}002a/{region_letter}{sid_lower}051a"
        
        # D10: /srv/nfs/{region_letter}{sid_lower}
        sheet.cell(row=10, column=4).value = f"/srv/nfs/{region_letter}{sid_lower}"
        sheet.cell(row=10, column=5).value = f"/srv/nfs/{region_letter}{sid_lower}002"
        
        # Process NFS directory structure (rows 15-18)
        nfs_mappings = [
            (15, f"<NFS Server>/srv/nfs/{region_letter}{sid_lower}/interface", f"/interface/sap{sid_lower}"),
            (16, f"<NFS Server>/srv/nfs/{region_letter}{sid_lower}/sapmnt", f"/sapmnt/{sid}"),
            (17, f"<NFS Server>/srv/nfs/{region_letter}{sid_lower}/{region_letter}{sid_lower}002", f"/app/{region_letter}{sid_lower}002"),
            (18, f"<NFS Server>/srv/nfs/{region_letter}{sid_lower}/trans", f"/usr/sap/trans")
        ]
        
        for row_num, col1_value, col4_value in nfs_mappings:
            sheet.cell(row=row_num, column=1).value = col1_value
            if col4_value:
                sheet.cell(row=row_num, column=4).value = col4_value

    def apply_db2_logic(sheet, sid, sid_lower, region_letter):
        """Apply DB2-specific cell updates"""
        # Update NFS Mounts
        sheet.cell(row=51, column=4).value = f"{region_letter}{sid_lower}000a:/srv/nfs/{region_letter}{sid_lower}/sapmnt"

    def apply_ascs_logic(sheet, sid, region, region_dr, afs_servername, afs_servername_dr, has_dr, is_production):
        """Apply ASCS-specific cell updates"""
        if has_dr:
            if is_production:
                # Production DR scenario
                sheet.cell(row=25, column=1).value = f"Primary - {region} ({afs_servername})"
                sheet.cell(row=27, column=1).value = f"Primary - {region} ({afs_servername})"
                sheet.cell(row=35, column=1).value = f"DR - {region_dr} ({afs_servername_dr}/)"
                sheet.cell(row=37, column=1).value = f"DR - {region_dr} ({afs_servername_dr}/)"
            else:
                # Non-production DR scenario
                sheet.cell(row=13, column=1).value = f"Primary - {region} ({afs_servername})"
                sheet.cell(row=15, column=1).value = f"Primary - {region} ({afs_servername})"
        else:
            # Non-DR scenario
            if not is_production:
                sheet.cell(row=13, column=1).value = f"Primary - {region} ({afs_servername})"
                sheet.cell(row=15, column=1).value = f"Primary - {region} ({afs_servername})"

    # =============================================================================
    # SHEET CONFIGURATION AND PROCESSING
    # =============================================================================
    
    # Define sheet mappings and their specific processing logic
    sheet_configs = {
        'nfs': {
            'condition': role_flags['has_nfs'],
            'sheet_mapping': {"ASCS+PAS+NFS": "single"},
            'target_name': "FS Layout ASCS+PAS+NFS",
            'custom_logic': lambda sheet: apply_nfs_logic(sheet, sid, sid_lower, region_letter, service_model_names),
            'use_version_logic': False
        },
        'hana': {
            'condition': role_flags['has_hana'],
            'base_role': "HANA",
            'target_name': "FS Layout HANA",
            'custom_logic': None,
            'use_version_logic': True
        },
        'db2': {
            'condition': role_flags['has_db2'],
            'base_role': "DB2 DB",
            'target_name': "FS Layout DB2",
            'custom_logic': lambda sheet: apply_db2_logic(sheet, sid, sid_lower, region_letter) if not is_production else None,
            'use_version_logic': True
        },
        'maxdb': {
            'condition': role_flags['has_maxdb'],
            'base_role': "MaxDB",
            'target_name': "FS Layout MaxDB",
            'custom_logic': None,
            'use_version_logic': True
        },
        'ascs': {
            'condition': role_flags['has_ascs'] or role_flags['has_ascs_dr'],
            'sheet_mapping': {"ASCS Non_Prod": "non_prod", "ASCS Prod": "prod"},
            'target_name': "FS Layout ASCS",
            'custom_logic': lambda sheet: apply_ascs_logic(sheet, sid, region, region_dr, afs_servername, afs_servername_dr, role_flags['has_ascs_dr'], is_production),
            'use_version_logic': False
        },
        'pas': {
            'condition': role_flags['only_pas_aas'],
            'sheet_mapping': {"PAS Non_Prod": "non_prod", "PAS Prod": "prod"},
            'target_name': "FS Layout PAS_AAS",
            'custom_logic': None,
            'use_version_logic': False
        },
        'apo': {
            'condition': role_flags['has_apo'],
            'base_role': "APO",
            'target_name': "FS Layout APO",
            'custom_logic': None,
            'use_version_logic': True
        },
        'iq': {
            'condition': role_flags['has_iq'],
            'base_role': "IQ",
            'target_name': "FS Layout IQ",
            'custom_logic': None,
            'use_version_logic': True
        },
        'optimizer': {
            'condition': role_flags['has_optimizer'],
            'base_role': "Optimizer",
            'target_name': "FS Layout Optimizer",
            'custom_logic': None,
            'use_version_logic': True
        }
    }

    # Process each sheet configuration
    for sheet_type, config in sheet_configs.items():
        if config['condition']:
            sheet = None
            
            # Use version-aware logic for roles that support versions
            if config.get('use_version_logic', False):
                sheet = select_and_rename_sheet_with_version(config['base_role'], config['target_name'])
            # Handle single sheets (no environment variants)
            elif 'sheet_mapping' in config and any(env == "single" for env in config['sheet_mapping'].values()):
                sheet_name = next(name for name, env in config['sheet_mapping'].items() if env == "single")
                # Check if sheet exists with case-insensitive comparison
                matching_sheet = next((name for name in workbook.sheetnames if name.lower() == sheet_name.lower()), None)
                if matching_sheet:
                    sheet = workbook[matching_sheet]
                    sheet.title = config['target_name']
                    sheet = workbook[config['target_name']]
            # Handle environment-based sheets (traditional logic)
            elif 'sheet_mapping' in config:
                sheet = select_and_rename_sheet(config['sheet_mapping'], config['target_name'])
            
            if sheet:
                # Apply custom logic if available
                if config['custom_logic']:
                    config['custom_logic'](sheet)
                
                # Always apply SID replacement
                replace_sid_in_sheet(sheet, sid)
        else:
            # Remove unused sheets
            if config.get('use_version_logic', False):
                # For version-aware sheets, remove all variations
                base_role = config['base_role']
                sheets_to_remove = []
                for sheet_name in workbook.sheetnames:
                    if base_role.lower() in sheet_name.lower() and ("prod" in sheet_name.lower() or "non_prod" in sheet_name.lower()):
                        sheets_to_remove.append(sheet_name)
                remove_sheets_if_unused(sheets_to_remove)
            elif 'sheet_mapping' in config:
                remove_sheets_if_unused(list(config['sheet_mapping'].keys()))

    print(f"Sheet processing completed for environment: {'Production' if is_production else 'Non-Production'}")
    return workbook