import pandas as pd
import json
import os
import openpyxl
from utils import generate_service_model_names, get_environment_code, get_instance_number, get_sap_region_letter, add_other_sheets, add_load_balancer_sheet

def process_non_prod_data_to_excel(json_file_path, template_path, output_path=None):
    """
    Process SAP form data from JSON and fill in the Excel template.
    
    Args:
        json_file_path (str): Path to the JSON file with form data
        template_path (str): Path to the Excel template
        output_path (str, optional): Path to save the output Excel file. Defaults to None.
    
    Returns:
        str: Path to the saved Excel file
    """
    # Check if files exist
    if not os.path.exists(json_file_path):
        raise FileNotFoundError(f"JSON file not found: {json_file_path}")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Excel template not found: {template_path}")
    
    # Track AAS server counts by city
    aas_counters = {"Amsterdam": 1, "Dublin": 1}
    
    # Load JSON data
    with open(json_file_path, 'r') as file:
        form_data = json.load(file)
    
    # Extract general config and server data
    general_config = form_data.get("general_config", {})
    server_data = form_data.get("server_data", [])
    
    # Load Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["SAP"]  # Access the SAP sheet
    
    # Start filling data from row 12
    start_row = 12
    
    # For each server, fill a row in the Excel sheet
    for i, server in enumerate(server_data):
        row = start_row + i
        
        # Fill in values using the correct column numbers from the provided list
        # Column 1: Azure Region
        sheet.cell(row=row, column=1).value = general_config.get("Azure Region", "")
        
        # Column 2: Environment
        sheet.cell(row=row, column=2).value = general_config.get("Environment", "")
        
        # Column 3: SID
        sid = general_config.get("SID", "")
        sheet.cell(row=row, column=3).value = sid
        
        # Column 4: Instance Number
        server_role = server.get("Server Role", "")
        environment = general_config.get("Environment", "")
        instance_number = get_instance_number(server_role, environment)
        sheet.cell(row=row, column=4).value = instance_number
        
        # Column 5: Server Role
        sheet.cell(row=row, column=5).value = server.get("Server Role", "")
        
        # Column 6: Service Criticality
        sheet.cell(row=row, column=6).value = general_config.get("Service Criticality", "")
        
        # Column 7: ITSG ID
        sheet.cell(row=row, column=7).value = general_config.get("ITSG ID", "")
        
        # Column 8: OS Version
        sheet.cell(row=row, column=8).value = server.get("OS Version", "")
        
        # Column 9: Licensing (default to "UL: BYOL")
        sheet.cell(row=row, column=9).value = "UL: BYOL"
        
        # Column 10: Azure Instance Type
        sheet.cell(row=row, column=10).value = server.get("Instance Type", "")
        
        # Column 11: Memory / CPU
        sheet.cell(row=row, column=11).value = server.get("Memory/CPU", "")
        
        # Column 12: Service Model name
        # Determine AAS counter if it's an AAS server
        aas_counter = None
        if "AAS" in server.get("Server Role", ""):
            # Extract city from Azure Region or use default
            city = "Amsterdam"  # Default
            if "Dublin" in general_config.get("Azure Region", ""):
                city = "Dublin"
            aas_counter = aas_counters[city]
            aas_counters[city] += 1
        
        service_model_name = generate_service_model_names(
            server.get("Server Role", ""), 
            general_config.get("SID", ""),
            general_config.get("SAP Region", "Sirius"),
            general_config.get("Azure Region").split("(")[1].split(")")[0],
            aas_counter=aas_counter
        )
        
        sheet.cell(row=row, column=12).value = service_model_name
        
        # Column 13: A Record / CNAME
        sheet.cell(row=row, column=13).value = general_config.get("Record Type", "")
        
        # Column 14: VNET (default to "STS Vnet")
        sheet.cell(row=row, column=14).value = "STS Vnet"
        
        # Column 15: Subnet/Zone
        sheet.cell(row=row, column=15).value = general_config.get("Subnet/Zone", "")
        
        # Column 16: Subnet (leave blank)
        sheet.cell(row=row, column=16).value = ""
        
        # Column 17: Subnet Group (leave blank)
        sheet.cell(row=row, column=17).value = ""
        
        # Column 18: Azure Subscription
        sheet.cell(row=row, column=18).value = general_config.get("Azure Subscription", "")
        
        # Column 19: Azure Resource Group Name (generate based on pattern)
        region_code = general_config.get("Azure Region Code", "").lower()
        environment_code = get_environment_code(general_config.get("Environment", ""))
        itsg_id = general_config.get("ITSG ID", "")
        sid = general_config.get("SID", "").upper()
        subscription = general_config.get("Azure Subscription", "")
        subscription_number = subscription.split("-")[1].split(" ")[0]
        
        # Format: <REGION_CODES>-sp01-s-<ITSG>-<SID>-IS01-rg
        resource_group_name = f"{region_code}-sp{subscription_number}-{environment_code}-{itsg_id}-{sid}-IS01-rg"
        sheet.cell(row=row, column=19).value = resource_group_name
        
        # Column 20: Cluster
        sheet.cell(row=row, column=20).value = server.get("Cluster", "")
        
        # Column 21: Proximity Placement Group (updated format)
        server_role = server.get("Server Role", "")
        if any(role in server_role for role in ["PAS", "AAS", "ASCS", "SCS"]):
            sap_region = general_config.get("SAP Region", "Sirius")
            region_letter = get_sap_region_letter(sap_region)
            az_zone = general_config.get("AZ Selection", "")

            # Format: <azure_region_code>-sp01-<SAPregion>-<ITSG>-<SID>-az<x>-ppg
            ppg = f"{region_code}-sp{subscription_number}-{region_letter}-{itsg_id}-{sid}-az{az_zone}-ppg"
            sheet.cell(row=row, column=21).value = ppg
        else:
            sheet.cell(row=row, column=21).value = ""
        
        # Column 22: Availability Set (updated format)
        if server.get("Availability Set", "").lower() == "yes":
            sap_region = general_config.get("SAP Region", "Sirius")
            region_letter = get_sap_region_letter(sap_region)
            az_zone = general_config.get("AZ Selection", "")
            
            # Format: <azure_region_code>-sp01-<SAPregion>-<ITSG>-<SID>-az<x>-as
            availability_set = f"{region_code}-sp{subscription_number}-{region_letter}-{itsg_id}-{sid}-az{az_zone}-as"
            sheet.cell(row=row, column=22).value = availability_set
        else:
            sheet.cell(row=row, column=22).value = ""
        
        # Column 23: Azure Availability Zone
        sheet.cell(row=row, column=23).value = general_config.get("AZ Selection", "")
        
        # Column 24: Accelerated Networking (default to "Yes")
        sheet.cell(row=row, column=24).value = "Yes"
        
        # Column 25: On Demand/Reservation
        sheet.cell(row=row, column=25).value = server.get("Reservation Type", "")
        
        # Column 26: Reservation Term
        sheet.cell(row=row, column=26).value = server.get("Reservation Term", "")
        
        # Column 27: OptInOptOut
        sheet.cell(row=row, column=27).value = server.get("OptInOptOut", "")
        
        # Column 28: Park My Cloud Schedule
        sheet.cell(row=row, column=28).value = server.get("Park My Cloud Schedule", "")
        
        # Column 29: Park My cloud team name and Member
        sheet.cell(row=row, column=29).value = server.get("Park My cloud team name and Member", "")
        
        # Columns 30-43: Type and Qty pairs (leaving blank)
        for col in range(30, 44):
            sheet.cell(row=row, column=col).value = ""
        
        # Column 44: Outbound Internet Access Required
        sheet.cell(row=row, column=44).value = server.get("Outbound Internet Access Required", "")
        
        # Column 45: Additional Requirements / Comments (leave blank)
        sheet.cell(row=row, column=45).value = ""
        
        # Column 46: Build RFC (leave blank)
        sheet.cell(row=row, column=46).value = ""
        
        # Column 47: iSCSI details (leave blank)
        sheet.cell(row=row, column=47).value = ""
        
        # Column 48: Instance Name (leave blank)
        sheet.cell(row=row, column=48).value = ""
        
        # Column 49: Private IP Address (leave blank)
        sheet.cell(row=row, column=49).value = ""
        
        # Column 50: TimeZone (default to "CET")=====
        sheet.cell(row=row, column=50).value = general_config.get("Timezone", "CET")
    
    
    # Remove Load Balancer Sheet if it exists
    if "Azure Load Balancer" in workbook.sheetnames:
        del workbook["Azure Load Balancer"]

    # Process the sheets after handling the SAP sheet
    add_other_sheets(json_file_path, template_path, workbook)
    
    # Save the workbook
    workbook.save(output_path)
    
    if output_path is None:
        output_path = f"Filled_{os.path.basename(template_path)}"
    
    return output_path

def fill_server_data(sheet, server, general_config, row, aas_counters):
    """
    Fill server data into a specific row of the Excel sheet.
    
    Args:
        sheet: Excel worksheet object
        server: Server data dictionary
        general_config: General configuration dictionary
        row: Row number to fill
    
    Returns:
        int: Next available row number
    """
    
    # Column 1: Azure Region
    if "DR" in server.get("Server Role"):
        region = "Azure: Northern Europe (Dublin) (IENO)" if "Amsterdam" in general_config.get("Azure Region") else "Azure: Western Europe (Amsterdam) (NLWE)"
        sheet.cell(row=row, column=1).value = region
    else:    
        sheet.cell(row=row, column=1).value = general_config.get("Azure Region", "")
    
    # Column 2: Environment
    if "DR" in server.get("Server Role"):
        sheet.cell(row=row, column=2).value = "Production-DR"
    else:    
        sheet.cell(row=row, column=2).value = general_config.get("Environment", "")
        
    # Column 3: SID
    sid = general_config.get("SID", "")
    sheet.cell(row=row, column=3).value = sid
    
    # Column 4: Instance Number
    server_role = server.get("Server Role", "")
    environment = general_config.get("Environment", "")
    instance_number = get_instance_number(server_role, environment)
    sheet.cell(row=row, column=4).value = instance_number
    
    # Column 5: Server Role
    sheet.cell(row=row, column=5).value = server.get("Server Role", "")
    
    # Column 6: Service Criticality
    sheet.cell(row=row, column=6).value = general_config.get("Service Criticality", "")
    
    # Column 7: ITSG ID
    sheet.cell(row=row, column=7).value = general_config.get("ITSG ID", "")
    
    # Column 8: OS Version
    sheet.cell(row=row, column=8).value = server.get("OS Version", "")
    
    # Column 9: Licensing (default to "UL: BYOL")
    sheet.cell(row=row, column=9).value = "UL: BYOL"
    
    # Column 10: Azure Instance Type
    # Handle both "Instance Type" and "Azure Instance Type" keys
    instance_type = server.get("Instance Type") or server.get("Azure Instance Type", "")
    sheet.cell(row=row, column=10).value = instance_type
    
    # Column 11: Memory / CPU
    sheet.cell(row=row, column=11).value = server.get("Memory/CPU", "")
    
    # Column 12: Service Model name
    azure_region = general_config.get("Azure Region", "")
    # Extract region code from the Azure Region string (e.g., "IENO" from "Azure: Northern Europe (Dublin) (IENO)")
    region_code = azure_region.split("(")[-1].replace(")", "") if "(" in azure_region else ""
    
    
    # Determine AAS counter if it's an AAS server
    aas_counter = None
    city = "Dublin"  # Default
    if "AAS" in server.get("Server Role", "") and "DR" not in server.get("Server Role"):
        # Extract city from Azure Region or use default
        if "Amsterdam" in general_config.get("Azure Region", ""):
            city = "Amsterdam"
        
    elif "AAS" in server.get("Server Role", "") and "DR" in server.get("Server Role"):
        if "Dublin" in general_config.get("Azure Region", ""):
            city = "Amsterdam"
        
        aas_counter = aas_counters[city]
        aas_counters[city] += 1
    
    service_model_name = generate_service_model_names(
        server.get("Server Role", ""), 
        general_config.get("SID", ""),
        general_config.get("SAP Region", "Sirius"),
        city,
        aas_counter=aas_counter
    )
    sheet.cell(row=row, column=12).value = service_model_name
    
    # Column 13: A Record / CNAME
    sheet.cell(row=row, column=13).value = server.get("Record Type", "")
    
    # Column 14: VNET (default to "STS Vnet")
    sheet.cell(row=row, column=14).value = "STS Vnet"
    
    # Column 15: Subnet/Zone
    sheet.cell(row=row, column=15).value = general_config.get("Subnet/Zone", "")
    
    # Column 16: Subnet (leave blank)
    sheet.cell(row=row, column=16).value = ""
    
    # Column 17: Subnet Group (leave blank)
    sheet.cell(row=row, column=17).value = ""
    
    # Column 18: Azure Subscription
    sheet.cell(row=row, column=18).value = general_config.get("Azure Subscription", "")
    
    # Column 19: Azure Resource Group Name (generate based on pattern)
    region_code = general_config.get("Azure Region Code", "").lower()
    environment_code = get_environment_code(general_config.get("Environment", ""))
    itsg_id = general_config.get("ITSG ID", "")
    sid = general_config.get("SID", "").upper()
    subscription = general_config.get("Azure Subscription", "")
    subscription_number = subscription.split("-")[1].split(" ")[0] if "-" in subscription else "01"
    
    # Format: <REGION_CODES>-sp01-s-<ITSG>-<SID>-IS01-rg
    resource_group_name = f"{region_code}-sp{subscription_number}-{environment_code}-{itsg_id}-{sid}-IS01-rg"
    sheet.cell(row=row, column=19).value = resource_group_name
    
    # Column 20: Cluster
    sheet.cell(row=row, column=20).value = server.get("Cluster", "")
    
    # Column 21: Proximity Placement Group (updated format)
    server_role = server.get("Server Role", "")
    if any(role in server_role for role in ["PAS", "AAS", "ASCS", "SCS"]):
        sap_region = general_config.get("SAP Region", "Sirius")
        region_letter = get_sap_region_letter(sap_region)
        if "-HA" in server_role:
            az_zone = server.get("HA_Zone","")
        else:
            az_zone = server.get("AZ Selection", "")

        # Format: <azure_region_code>-sp01-<SAPregion>-<ITSG>-<SID>-az<x>-ppg
        ppg = f"{region_code}-sp{subscription_number}-{region_letter}-{itsg_id}-{sid}-az{az_zone}-ppg"
        sheet.cell(row=row, column=21).value = ppg
    else:
        sheet.cell(row=row, column=21).value = ""
    
    # Column 22: Availability Set (updated format)
    if server.get("Availability Set", "").lower() == "yes":
        sap_region = general_config.get("SAP Region", "Sirius")
        region_letter = get_sap_region_letter(sap_region)
        if "-HA" in server_role:
            az_zone = server.get("HA_Zone","")
        else:
            az_zone = server.get("AZ Selection", "")
            
        # Format: <azure_region_code>-sp01-<SAPregion>-<ITSG>-<SID>-az<x>-as
        availability_set = f"{region_code}-sp{subscription_number}-{region_letter}-{itsg_id}-{sid}-az{az_zone}-as"
        sheet.cell(row=row, column=22).value = availability_set
    else:
        sheet.cell(row=row, column=22).value = ""
    
    # Column 23: Azure Availability Zone
    if "-HA" in server_role:
        az_zone = server.get("HA_Zone","")
    else:
        az_zone = server.get("AZ Selection", "")
    
    sheet.cell(row=row, column=23).value = az_zone

    # Column 24: Accelerated Networking (default to "Yes")
    sheet.cell(row=row, column=24).value = "Yes"
    
    # Column 25: On Demand/Reservation
    sheet.cell(row=row, column=25).value = server.get("Reservation Type", "")
    
    # Column 26: Reservation Term
    sheet.cell(row=row, column=26).value = server.get("Reservation Term", "")
    
    # Column 27: OptInOptOut
    sheet.cell(row=row, column=27).value = server.get("OptInOptOut", "")
    
    # Column 28: Park My Cloud Schedule
    sheet.cell(row=row, column=28).value = server.get("Park My Cloud Schedule", "")
    
    # Column 29: Park My cloud team name and Member
    sheet.cell(row=row, column=29).value = server.get("Park My cloud team name and Member", "")
    
    # Columns 30-43: Type and Qty pairs (leaving blank)
    for col in range(30, 44):
        sheet.cell(row=row, column=col).value = ""
    
    # Column 44: Outbound Internet Access Required
    sheet.cell(row=row, column=44).value = server.get("Outbound Internet Access Required", "")
    
    # Column 45: Additional Requirements / Comments (leave blank)
    sheet.cell(row=row, column=45).value = ""
    
    # Column 46: Build RFC (leave blank)
    sheet.cell(row=row, column=46).value = ""
    
    # Column 47: iSCSI details (leave blank)
    sheet.cell(row=row, column=47).value = ""
    
    # Column 48: Instance Name (leave blank)
    sheet.cell(row=row, column=48).value = ""
    
    # Column 49: Private IP Address (leave blank)
    sheet.cell(row=row, column=49).value = ""
    
    # Column 50: TimeZone (default to "CET")
    sheet.cell(row=row, column=50).value = general_config.get("Timezone", "CET")
    
    return row + 1

def process_prod_data_to_excel(json_file_path, template_path, output_path=None):
    """
    Process SAP form data from JSON for Production environment and fill in the Excel template.
    
    Args:
        json_file_path (str): Path to the JSON file with form data
        template_path (str): Path to the Excel template
        output_path (str, optional): Path to save the output Excel file. Defaults to None.
    
    Returns:
        str: Path to the saved Excel file
    """
    # Check if files exist
    if not os.path.exists(json_file_path):
        raise FileNotFoundError(f"JSON file not found: {json_file_path}")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Excel template not found: {template_path}")
    
    # Track AAS server counts by city
    aas_counters = {"Amsterdam": 1, "Dublin": 1}
    
    # Load JSON data
    with open(json_file_path, 'r') as file:
        form_data = json.load(file)
        
    # Extract general config and server data
    general_config = form_data.get("general_config", {})
    primary_servers = form_data.get("primary_servers", [])
    dr_servers = form_data.get("dr_servers", [])
    
    # Load Excel template
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["SAP"]  # Access the SAP sheet
    
    # Start filling data from row 12
    current_row = 12
    
    # Process Primary servers first
    for server in primary_servers:
        # Fill primary server data
        current_row = fill_server_data(sheet, server, general_config, current_row, aas_counters)
        
        # Check if server has cluster (HA counterpart)
        if server.get("Cluster", "").lower() == "yes" and server.get("HA_Role"):
            # Create HA server entry with same config but different role
            ha_server = server.copy()
            ha_server["Server Role"] = server.get("HA_Role")
            current_row = fill_server_data(sheet, ha_server, general_config, current_row, aas_counters)
    
    # Leave 1 blank row
    current_row += 1  # First blank row
    
    # Fill the next row with blue color (columns A to AX = 1 to 50)
    for col in range(1, 51):
        cell = sheet.cell(row=current_row, column=col)
        cell.fill = openpyxl.styles.PatternFill(start_color="657C9C", end_color="657C9C", fill_type="solid")
    
    current_row += 2  # Skip the blue row and add another blank row
    
    # Process DR servers
    for server in dr_servers:
        current_row = fill_server_data(sheet, server, general_config, current_row, aas_counters)
    
    # Add Load Balancer Sheet
    add_load_balancer_sheet(workbook, general_config, primary_servers)
    
    # Process the sheets after handling the SAP sheet
    add_other_sheets(json_file_path, template_path, workbook)
    
    # Save the workbook
    if output_path is None:
        output_path = f"Filled_{os.path.basename(template_path)}"
    
    workbook.save(output_path)
    
    return output_path


if __name__ == "__main__":
    # This block only executes when the script is run directly
    # It won't run when the script is imported
    # import sys
    
    # # Check if command line arguments are provided
    # if len(sys.argv) < 3:
    #     print("Usage: python excel_processor.py <json_file> <template_file> [output_file]")
    #     sys.exit(1)
    
    # json_file = sys.argv[1]
    # template_file = sys.argv[2]
    # output_file = sys.argv[3] if len(sys.argv) > 3 else None
    json_file = "sap_form_data_prod_E4E.json"
    template_file = "Template.xlsx"
    output_file = "Filled_SAP_Template.xlsx"
    
    try:
        # output_path = process_non_prod_data_to_excel(json_file, template_file, output_file)
        output_path = process_prod_data_to_excel(json_file, template_file, output_file)
        print(f"Processing completed successfully. Output saved to: {output_path}")
    except Exception as e:
        print(f"Error: {str(e)}")
        # sys.exit(1)